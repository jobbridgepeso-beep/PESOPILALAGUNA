"""
Excel export service for the JobBridge application.

Uses OpenPyXL to generate formatted Excel workbooks.

Requirements: 10.6, 11.6, 13.2
"""

from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Header style
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def generate_attendance_excel(attendance_data: list[dict[str, Any]]) -> bytes:
    """Generate a job fair attendance Excel report.

    Args:
        attendance_data: List of attendance record dicts.

    Returns:
        Excel file bytes (.xlsx).

    Requirements: 10.6
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["#", "Name", "Email", "Phone", "Scanned At", "Scanned By"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for i, record in enumerate(attendance_data, start=1):
        ws.append([
            i,
            record.get("name", ""),
            record.get("email", ""),
            record.get("phone", ""),
            str(record.get("scanned_at", "")),
            record.get("scanned_by", ""),
        ])

    _auto_width(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_lmi_report_excel(report_data: dict[str, Any]) -> bytes:
    """Generate an LMI report Excel workbook with multiple worksheets.

    Args:
        report_data: Dict containing LMI statistics.

    Returns:
        Excel file bytes (.xlsx).

    Requirements: 13.2
    """
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    _style_header_row(ws_summary, len(summary_headers))

    stats = report_data.get("stats", {})
    for key, value in stats.items():
        ws_summary.append([key.replace("_", " ").title(), value])
    _auto_width(ws_summary)

    # Top industries sheet
    if "top_industries" in report_data:
        ws_ind = wb.create_sheet("Top Industries")
        ws_ind.append(["Industry", "Count"])
        _style_header_row(ws_ind, 2)
        for item in report_data["top_industries"]:
            ws_ind.append([item.get("industry", ""), item.get("count", 0)])
        _auto_width(ws_ind)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_program_enrollment_excel(enrollment_data: list[dict[str, Any]]) -> bytes:
    """Generate a program enrollment Excel report.

    Args:
        enrollment_data: List of program application dicts.

    Returns:
        Excel file bytes (.xlsx).

    Requirements: 11.6
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollments"

    headers = ["#", "Name", "Email", "Program", "Status", "Applied At", "Decision"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for i, record in enumerate(enrollment_data, start=1):
        ws.append([
            i,
            record.get("name", ""),
            record.get("email", ""),
            record.get("program_type", "").upper(),
            record.get("status", "").title(),
            str(record.get("created_at", "")),
            record.get("decision_reason", ""),
        ])

    _auto_width(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
